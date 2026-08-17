from __future__ import annotations

import csv
import re
import textwrap
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable

from PIL import Image, ImageDraw, ImageFont, ImageOps, ImageStat

from premium_prompt import PRODUCTO_EN_USO_PREMIUM_PROMPT as _PRODUCTO_EN_USO_PREMIUM_PROMPT_RAW
from tonka_prompt import CATALOGO_A4_TONKA_PROMPT


def _repair_attachment_encoding(value: str) -> str:
    """Repair only mojibake sequences while preserving characters already decoded."""
    pattern = re.compile(r"Ã.|Â.|â..")

    def repair(match: re.Match[str]) -> str:
        try:
            return match.group(0).encode("cp1252").decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            return match.group(0)

    return pattern.sub(repair, value)


PRODUCTO_EN_USO_PREMIUM_PROMPT = _repair_attachment_encoding(_PRODUCTO_EN_USO_PREMIUM_PROMPT_RAW)


SUPPORTED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
GROUPING_OPTIONS = ("Individual", "Por prefijo/SKU", "Por subcarpeta")
QUALITY_OPTIONS = ("Baja", "Media", "Alta")
OUTPUT_STYLE_OPTIONS = ("Imagen IA", "Infografia con texto exacto")
OUTPUT_FORMAT_OPTIONS = ("PNG", "JPG")

QUALITY_API_VALUES = {
    "Baja": "low",
    "Media": "medium",
    "Alta": "high",
}

# Official per-image output estimates. Image/text input tokens are additional.
OUTPUT_COST_USD = {
    "gpt-image-1-mini": {
        "low": {"1024x1024": 0.005, "1024x1536": 0.006, "1536x1024": 0.006},
        "medium": {"1024x1024": 0.011, "1024x1536": 0.015, "1536x1024": 0.015},
        "high": {"1024x1024": 0.036, "1024x1536": 0.052, "1536x1024": 0.052},
    },
    "gpt-image-2": {
        "low": {"1024x1024": 0.006, "1024x1536": 0.005, "1536x1024": 0.005},
        "medium": {"1024x1024": 0.053, "1024x1536": 0.041, "1536x1024": 0.041},
        "high": {"1024x1024": 0.211, "1024x1536": 0.165, "1536x1024": 0.165},
    },
    "gpt-image-2-2026-04-21": {
        "low": {"1024x1024": 0.006, "1024x1536": 0.005, "1536x1024": 0.005},
        "medium": {"1024x1024": 0.053, "1024x1536": 0.041, "1536x1024": 0.041},
        "high": {"1024x1024": 0.211, "1024x1536": 0.165, "1536x1024": 0.165},
    },
}

PROMPT_PRESETS = {
    "Personalizado": "",
    "Producto en uso premium": PRODUCTO_EN_USO_PREMIUM_PROMPT,
    "Catálogo A4 Tonka": CATALOGO_A4_TONKA_PROMPT,
    "Catalogo blanco": (
        "Extrae y conserva exactamente el producto de la imagen. Colocalo completo y centrado sobre "
        "un fondo blanco opaco, con silueta limpia, colores fieles y una sombra de contacto sutil. "
        "No cambies geometria, marca, etiquetas, texto, piezas ni materiales. No agregues accesorios, "
        "texto publicitario, logotipos nuevos ni marcas de agua. Fotografia comercial fotorrealista."
    ),
    "Escena de uso": (
        "Identifica el producto principal y conserva exactamente su forma, colores, marca, piezas y "
        "proporciones. Presentalo como protagonista en una escena profesional y realista relacionada "
        "con su uso evidente. Si el uso no es claro, utiliza estudio blanco. No inventes funciones, "
        "accesorios, especificaciones, texto ni variantes. Iluminacion comercial natural."
    ),
    "Infografia comercial": (
        "Crea una fotografia publicitaria limpia del producto para usar como base de una infografia. "
        "Conserva exactamente geometria, colores, marca, etiquetas, piezas y proporciones. Coloca el "
        "producto completo sobre fondo blanco o escenario sutil de uso, dejando espacio negativo a la "
        "izquierda para texto que se agregara despues. No generes textos, iconos ni marcas de agua."
    ),
}


@dataclass
class ProductJob:
    key: str
    images: list[Path]
    metadata: dict[str, str] = field(default_factory=dict)

    @property
    def primary_image(self) -> Path:
        return self.images[0]


@dataclass
class ReportRow:
    key: str
    source_files: str
    output_file: str
    status: str
    error: str = ""
    model: str = ""
    quality: str = ""
    size: str = ""
    estimated_cost_usd: float = 0.0
    elapsed_seconds: float = 0.0


def discover_images(folder: Path, recursive: bool = False) -> list[Path]:
    iterator = folder.rglob("*") if recursive else folder.iterdir()
    return sorted(
        path
        for path in iterator
        if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS
    )


def normalize_product_key(stem: str) -> str:
    value = stem.strip()
    view_words = (
        "frente|frontal|atras|reverso|lateral|detalle|empaque|caja|uso|vista|"
        "front|back|side|detail|package"
    )
    value = re.sub(rf"(?:[_ -](?:{view_words}))+$", "", value, flags=re.IGNORECASE)
    return value or stem


def _normalize_metadata_key(value: str) -> str:
    return normalize_product_key(Path(value.strip()).stem).casefold()


def load_metadata(path: Path | None) -> dict[str, dict[str, str]]:
    if not path:
        return {}
    if not path.is_file():
        raise ValueError(f"No existe el archivo de datos: {path}")

    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(4096)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            except csv.Error:
                dialect = csv.excel
            rows = list(csv.DictReader(handle, dialect=dialect))
    elif suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Instala openpyxl para leer archivos Excel.") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(values, [])]
        rows = [dict(zip(headers, ("" if value is None else str(value) for value in row))) for row in values]
        workbook.close()
    else:
        raise ValueError("El archivo de datos debe ser CSV o XLSX.")

    result: dict[str, dict[str, str]] = {}
    key_candidates = ("codigo", "código", "sku", "archivo", "file", "producto_id")
    for raw_row in rows:
        row = {str(key).strip().casefold(): str(value or "").strip() for key, value in raw_row.items() if key}
        raw_key = next((row[name] for name in key_candidates if row.get(name)), "")
        if raw_key:
            result[_normalize_metadata_key(raw_key)] = row
    return result


def build_product_jobs(
    folder: Path,
    grouping: str = "Individual",
    metadata_path: Path | None = None,
) -> list[ProductJob]:
    recursive = grouping == "Por subcarpeta"
    images = discover_images(folder, recursive=recursive)
    metadata = load_metadata(metadata_path)
    groups: dict[str, list[Path]] = {}

    for image in images:
        if grouping == "Por prefijo/SKU":
            key = normalize_product_key(image.stem)
        elif grouping == "Por subcarpeta":
            key = image.parent.name
        else:
            key = image.stem
        groups.setdefault(key, []).append(image)

    jobs = []
    for key, references in sorted(groups.items(), key=lambda item: item[0].casefold()):
        row = metadata.get(_normalize_metadata_key(key), {})
        if not row:
            for reference in references:
                row = metadata.get(_normalize_metadata_key(reference.stem), {})
                if row:
                    break
        jobs.append(ProductJob(key=key, images=references, metadata=row))
    return jobs


def metadata_value(metadata: dict[str, str], *names: str, default: str = "") -> str:
    for name in names:
        value = metadata.get(name.casefold(), "").strip()
        if value:
            return value
    return default


def metadata_benefits(metadata: dict[str, str], limit: int = 5) -> list[str]:
    benefits = []
    for index in range(1, limit + 1):
        value = metadata_value(metadata, f"beneficio{index}", f"beneficio {index}", f"benefit{index}")
        if value:
            benefits.append(value)
    return benefits


def metadata_prompt_facts(metadata: dict[str, str], default_name: str) -> list[str]:
    """Convert every non-empty spreadsheet cell into a verified prompt fact."""
    groups = (
        ("Código/SKU", ("codigo", "código", "sku", "archivo", "file", "producto_id")),
        ("Producto", ("producto", "nombre", "name")),
        ("Marca exacta", ("marca", "brand")),
        ("Descripción", ("descripcion", "descripción", "subtitulo", "subtítulo", "subtitle")),
        ("Tamaño", ("tamano", "tamaño", "size", "talla")),
        ("Piezas/cantidad", ("piezas", "pieza", "pz", "pcs", "cantidad", "unidades")),
        ("Medidas", ("medidas", "medida", "dimensiones", "dimensions")),
        ("Material", ("material", "materiales")),
        ("Edad recomendada", ("edad", "edad recomendada", "rango de edad")),
        ("Peso", ("peso", "peso producto", "peso del producto")),
        ("Peso máximo", ("peso maximo", "peso máximo", "peso soportado", "capacidad de peso")),
        ("Capacidad", ("capacidad",)),
        ("Modelo", ("modelo", "model")),
        ("Color", ("color", "colores")),
    )
    facts: list[str] = []
    used: set[str] = set()
    for label, aliases in groups:
        for alias in aliases:
            value = metadata.get(alias.casefold(), "").strip()
            if value:
                facts.append(f"{label}: {value}")
                used.update(item.casefold() for item in aliases)
                break

    if not any(fact.startswith("Producto:") for fact in facts):
        facts.append(f"Producto: {default_name}")

    for index in range(1, 6):
        aliases = (f"beneficio{index}", f"beneficio {index}", f"benefit{index}")
        value = metadata_value(metadata, *aliases)
        used.update(alias.casefold() for alias in aliases)
        if value:
            facts.append(f"Beneficio verificado {index}: {value}")

    used.update({"pie", "footer"})
    footer = metadata_value(metadata, "pie", "footer")
    if footer:
        facts.append(f"Texto inferior: {footer}")

    for key, value in metadata.items():
        clean_value = value.strip()
        if clean_value and key.casefold() not in used:
            label = re.sub(r"[_-]+", " ", key).strip().capitalize()
            facts.append(f"{label}: {clean_value}")
    return facts


def enrich_prompt(base_prompt: str, job: ProductJob) -> str:
    if not job.metadata:
        return base_prompt
    facts = metadata_prompt_facts(job.metadata, job.key)
    return (
        f"{base_prompt.strip()}\n\nDATOS VERIFICADOS DEL EXCEL. Respeta exactamente estos valores; "
        "no los cambies, completes ni inventes información adicional:\n- "
        + "\n- ".join(facts)
    ).strip()


def estimate_output_cost(model: str, quality: str, size: str, fallback: float) -> float:
    return OUTPUT_COST_USD.get(model, {}).get(quality, {}).get(size, fallback)


def choose_output_path(output_folder: Path, key: str, version_existing: bool, output_format: str = "PNG") -> Path:
    safe_key = re.sub(r'[<>:"/\\|?*]+', "_", key).strip(" .") or "producto"
    extension = ".jpg" if output_format.upper() == "JPG" else ".png"
    base = output_folder / f"{safe_key}{extension}"
    if not base.exists() or not version_existing:
        return base
    version = 2
    while True:
        candidate = output_folder / f"{safe_key}_v{version}{extension}"
        if not candidate.exists():
            return candidate
        version += 1


def _font(size: int, bold: bool = False):
    candidates = [
        Path("C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return ImageFont.truetype(str(candidate), size=size)
    return ImageFont.load_default()


def _draw_wrapped(draw: ImageDraw.ImageDraw, text: str, box: tuple[int, int, int, int], font, fill, spacing=6):
    x1, y1, x2, y2 = box
    approximate_chars = max(12, int((x2 - x1) / max(8, getattr(font, "size", 20) * 0.55)))
    lines = textwrap.wrap(text, width=approximate_chars)
    y = y1
    for line in lines:
        bbox = draw.textbbox((x1, y), line, font=font)
        height = bbox[3] - bbox[1]
        if y + height > y2:
            break
        draw.text((x1, y), line, font=font, fill=fill)
        y += height + spacing
    return y


def compose_infographic(ai_image_path: Path, output_path: Path, job: ProductJob) -> None:
    canvas = Image.new("RGB", (1080, 1080), "#f7f5f1")
    with Image.open(ai_image_path) as source:
        hero = ImageOps.contain(ImageOps.exif_transpose(source).convert("RGB"), (680, 820), Image.Resampling.LANCZOS)
    hero_x = 1080 - hero.width - 28
    hero_y = 80 + (820 - hero.height) // 2
    canvas.paste(hero, (hero_x, hero_y))

    draw = ImageDraw.Draw(canvas)
    navy, red, gray = "#17365d", "#c62032", "#4b5563"
    metadata = job.metadata
    name = metadata_value(metadata, "producto", "nombre", "name", default=job.key.replace("_", " "))
    brand = metadata_value(metadata, "marca", "brand")
    subtitle = metadata_value(
        metadata, "descripcion", "descripción", "subtitulo", "subtítulo", "subtitle"
    )
    benefits = metadata_benefits(metadata)

    if brand:
        draw.text((46, 40), brand, font=_font(36, True), fill=red)
    title_y = 92 if brand else 45
    title_end = _draw_wrapped(draw, name.upper(), (46, title_y, 390, 250), _font(50, True), navy, 4)
    if subtitle:
        _draw_wrapped(draw, subtitle, (46, title_end + 12, 390, 320), _font(24), gray, 4)

    draw.line((46, 330, 365, 330), fill=red, width=4)
    benefit_y = 365
    if benefits:
        for benefit in benefits[:5]:
            draw.ellipse((48, benefit_y + 2, 78, benefit_y + 32), fill=navy)
            draw.text((58, benefit_y + 4), "✓", font=_font(18, True), fill="white")
            benefit_y = _draw_wrapped(
                draw, benefit, (92, benefit_y, 390, benefit_y + 100), _font(24, True), navy, 3
            ) + 24
    else:
        draw.text((46, benefit_y), "PRESENTACION PROFESIONAL", font=_font(22, True), fill=navy)
        draw.text((46, benefit_y + 38), "Detalles fieles al producto original", font=_font(19), fill=gray)

    draw.rounded_rectangle((35, 955, 1045, 1040), radius=24, fill=navy)
    footer = metadata_value(metadata, "pie", "footer", default="IMAGEN DE CATALOGO · LISTA PARA REVISION")
    _draw_wrapped(draw, footer, (65, 978, 1015, 1028), _font(24, True), "white", 2)
    if output_path.suffix.lower() in {".jpg", ".jpeg"}:
        canvas.save(output_path, "JPEG", quality=95, subsampling=0, optimize=True)
    else:
        canvas.save(output_path, "PNG")


def validate_output(path: Path, expected_size: tuple[int, int] | None = None) -> list[str]:
    issues = []
    if not path.exists() or path.stat().st_size == 0:
        return ["El archivo de salida no existe o esta vacio."]
    try:
        with Image.open(path) as image:
            image.verify()
        with Image.open(path) as image:
            if expected_size and image.size != expected_size:
                issues.append(f"Tamano inesperado: {image.size[0]}x{image.size[1]}")
            if image.mode not in {"RGB", "RGBA"}:
                issues.append(f"Modo de color inusual: {image.mode}")
            sample = ImageOps.exif_transpose(image).convert("RGB").resize((64, 64))
            variance = sum(ImageStat.Stat(sample).var) / 3
            if variance < 4:
                issues.append("La imagen parece vacia o con contraste extremadamente bajo.")
    except Exception as exc:
        issues.append(f"PNG invalido: {exc}")
    return issues


def write_report(output_folder: Path, rows: Iterable[ReportRow]) -> Path:
    report_path = output_folder / f"reporte_lote_{datetime.now():%Y%m%d_%H%M%S}.csv"
    fieldnames = list(ReportRow.__dataclass_fields__)
    with report_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            data = {name: getattr(row, name) for name in fieldnames}
            data["estimated_cost_usd"] = f"{row.estimated_cost_usd:.4f}"
            data["elapsed_seconds"] = f"{row.elapsed_seconds:.2f}"
            writer.writerow(data)
    return report_path


def create_metadata_template(path: Path) -> None:
    headers = [
        "codigo", "producto", "marca", "descripcion", "tamano", "piezas", "medidas", "material",
        "edad", "peso", "capacidad", "modelo", "color", "beneficio1", "beneficio2", "beneficio3",
        "beneficio4", "pie"
    ]
    example = [
        "SKU-001", "Nombre del producto", "Marca", "Descripcion completa", "Tamaño", "4 pz",
        "Medidas", "Material", "Edad", "Peso", "Capacidad", "Modelo", "Color", "Beneficio 1",
        "Beneficio 2", "", "", "Texto inferior"
    ]
    if path.suffix.lower() == ".xlsx":
        from openpyxl import Workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Productos"
        sheet.append(headers)
        sheet.append(example)
        workbook.save(path)
    else:
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            writer.writerow(example)
